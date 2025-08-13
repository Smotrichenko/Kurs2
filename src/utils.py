import csv
import json
import logging
import os
from json import JSONDecodeError
from typing import Any, Dict, List

from openpyxl import load_workbook

from src.external_api import convert_to_rub

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

log_file_path = os.path.join(LOG_DIR, "utils.log")
logger = logging.getLogger("utils")
logger.setLevel(logging.DEBUG)
file_handler = logging.FileHandler(log_file_path, mode="w", encoding="utf=8")
file_formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s: %(message)s")
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)


def load_transactions(filename: str) -> List[Dict[str, Any]]:
    """
    Загружает транзакции из JSON, CSV или XLSX файла.
    Возвращает список словарей.
    """
    if not os.path.exists(filename):
        logger.error(f"Файл {filename} не найден.")
        print(f"Файл по пути {filename} не найден")
        return []

    ext = os.path.splitext(filename)[1].lower()
    logger.debug(f"Определён формат файла: {ext}")

    if ext == ".json":
        return load_from_json(filename)
    elif ext == ".csv":
        return load_from_csv(filename)
    elif ext == ".xlsx":
        return load_from_xlsx(filename)
    else:
        logger.error(f"Неизвестный формат файла: {ext}")
        print(f"Неизвестный формат файла: {ext}")
        return []


def load_from_json(filename: str) -> List[Dict[str, Any]]:
    try:
        logger.debug("Открываем файл JSON")
        with open(filename, "r", encoding="utf-8") as file:
            data = json.load(file)
            if not isinstance(data, List):
                print("Ожидается список транзакций в JSON-файле")
                return []
            return data

    except FileNotFoundError as ex:
        logger.error(f"Произошла ошибка: {ex}")
        print(f"Файл по пути {filename} не найден")
        return []
    except JSONDecodeError as ex:
        logger.error(f"Произошла ошибка: {ex}")
        print(f"Ошибка чтения JSON в файле {filename}")
        return []


def load_from_csv(filename: str) -> List[Dict[str, Any]]:
    logger.debug("Открываем CSV файл")
    transactions = []
    with open(filename, "r", encoding="utf-8") as file:
        reader = csv.DictReader(file, delimiter=";")
        for row in reader:
            transaction = {
                "id": row.get("id"),
                "state": row.get("state", "").upper(),
                "date": row.get("date"),
                "operationAmount": {
                    "amount": row.get("amount"),
                    "currency": {"code": row.get("currency_code", "RUB")},
                },
                "description": row.get("description", ""),
                "from": row.get("from"),
                "to": row.get("to"),
            }

            transactions.append(transaction)
    return transactions


def load_from_xlsx(filename: str) -> List[Dict[str, Any]]:
    logger.debug("Открываем XLSX файл")
    transactions = []
    workbook = load_workbook(filename=filename, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        transaction = {
            "id": row_data.get("id"),
            "state": row_data.get("state"),
            "date": row_data.get("date"),
            "from": row_data.get("from"),
            "to": row_data.get("to"),
            "description": row_data.get("description"),
            "operationAmount": {
                "amount": row_data.get("amount"),
                "currency": {"code": row_data.get("currency_code", "RUB")},
            },
        }
        transactions.append(transaction)
    return transactions


def get_transaction_amount(transactions: Dict[str, Any]) -> float:
    """
    Принимает транзакцию и возвращает сумму в рублях.
    Если валюта не рубль — конвертирует через внешний API.
    """
    try:
        logger.debug(f'Принимаем транзакцию {transactions.get("id")}')
        operation_amount = transactions.get("operationAmount")
        if not operation_amount or not isinstance(operation_amount, dict):
            return 0.0
        currency = operation_amount.get("currency", {}).get("currency_code", "RUB")
        if currency == "RUB":
            return float(operation_amount.get("amount", 0.0))

        logger.debug("Конвентируем через внешний API")
        return convert_to_rub(operation_amount)

    except Exception as ex:
        logger.error(f"Произошла ошибка: {ex}")
        return 0.0
